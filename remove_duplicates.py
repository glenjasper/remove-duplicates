#!/usr/bin/env python
# -*- coding: utf-8 -*-
import os
import sys
import time
import argparse
import traceback
import xlsxwriter
from tqdm import tqdm
from openpyxl import load_workbook
from crossref.restful import Works
from colorama import init
init()

def menu(args):
    parser = argparse.ArgumentParser(description = "This script eliminates the duplicated records from formatted .xlsx files from Scopus, Web of Science, PubMed, PubMed Central or Dimensions. Is mandatory that there be at least 2 different files from 2 different databases.", epilog = "Thank you!")
    parser.add_argument("-f", "--files", required = True, help = ".xlsx files separated by comma")
    parser.add_argument("-o", "--output", help = "Output folder")
    parser.add_argument("--version", action = "version", version = "%s %s" % ('%(prog)s', orr.VERSION))
    args = parser.parse_args()

    orr.INPUT_XLS_FILES = args.files
    file_list = orr.INPUT_XLS_FILES.split(',')
    for file in file_list:
        file_name = os.path.basename(file)
        file_path = os.path.dirname(file)
        if file_path is None or file_path == "":
            file_path = os.getcwd().strip()

        this_file = os.path.join(file_path, file_name)
        if not orr.check_path(this_file):
            orr.show_print("%s: error: the file '%s' doesn't exist" % (os.path.basename(__file__), this_file), showdate = False, font = oscihub.YELLOW)
            orr.show_print("%s: error: the following arguments are required: -f/--files" % os.path.basename(__file__), showdate = False, font = oscihub.YELLOW)
            exit()

        if os.path.basename(this_file) == orr.NAME_XLS_FILE_SCOPUS:
            orr.XLS_FILE_SCOPUS = this_file
        elif os.path.basename(this_file) == orr.NAME_XLS_FILE_WOS:
            orr.XLS_FILE_WOS = this_file
        elif os.path.basename(this_file) == orr.NAME_XLS_FILE_PUBMED:
            orr.XLS_FILE_PUBMED = this_file
        elif os.path.basename(this_file) == orr.NAME_XLS_FILE_PUBMED_CENTRAL:
            orr.XLS_FILE_PUBMED_CENTRAL = this_file
        elif os.path.basename(this_file) == orr.NAME_XLS_FILE_DIMENSIONS:
            orr.XLS_FILE_DIMENSIONS = this_file

    if args.output:
        output_name = os.path.basename(args.output)
        output_path = os.path.dirname(args.output)
        if output_path is None or output_path == "":
            output_path = os.getcwd().strip()

        orr.OUTPUT_PATH = os.path.join(output_path, output_name)
        created = orr.create_directory(orr.OUTPUT_PATH)
        if not created:
            orr.show_print("%s: error: Couldn't create folder '%s'" % (os.path.basename(__file__), orr.OUTPUT_PATH), showdate = False, font = oscihub.YELLOW)
            exit()
    else:
        orr.OUTPUT_PATH = os.getcwd().strip()
        orr.OUTPUT_PATH = os.path.join(orr.OUTPUT_PATH, 'output_remove_duplicate')
        orr.create_directory(orr.OUTPUT_PATH)

class RemoveDuplicate:

    def __init__(self):
        self.VERSION = 1.0

        self.INPUT_XLS_FILES = None
        self.OUTPUT_PATH = None
        self.DICT_XLS_FILES = {}

        self.ROOT_DIR = os.path.dirname(os.path.realpath(__file__))
        self.LOG_NAME = "run_%s_%s.log" % (os.path.splitext(os.path.basename(__file__))[0], time.strftime('%Y%m%d'))
        self.LOG_FILE = None

        # Repositories
        self.REPOSITORY_SCOPUS = "Scopus"
        self.REPOSITORY_WOS = "Web of Science"
        self.REPOSITORY_PUBMED = "PubMed"
        self.REPOSITORY_PUBMED_CENTRAL = "PubMed Central"
        self.REPOSITORY_DIMENSIONS = "Dimensions"

        # Xls Summary
        # Input
        self.XLS_FILE_SCOPUS = None
        self.XLS_FILE_WOS = None
        self.XLS_FILE_PUBMED = None
        self.XLS_FILE_PUBMED_CENTRAL = None
        self.XLS_FILE_DIMENSIONS = None
        self.NAME_XLS_FILE_SCOPUS = 'input_scopus.xlsx'
        self.NAME_XLS_FILE_WOS = 'input_wos.xlsx'
        self.NAME_XLS_FILE_PUBMED = 'input_pubmed.xlsx'
        self.NAME_XLS_FILE_PUBMED_CENTRAL = 'input_pmc.xlsx'
        self.NAME_XLS_FILE_DIMENSIONS = 'input_dimensions.xlsx'
        # Output
        self.XLS_FILE_OUTPUT = 'summary_unique_dois.xlsx'
        self.XLS_SHEET_DETAIL = 'Detail'
        self.XLS_SHEET_DUPLICATES = 'Duplicates'

        # Xls Columns
        self.xls_col_item = 'Item'
        self.xls_col_title = 'Title'
        self.xls_col_year = 'Year'
        self.xls_col_doi = 'DOI'
        self.xls_col_document_type = 'Document Type'
        self.xls_col_languaje = 'Language'
        self.xls_col_cited_by = 'Cited By'
        self.xls_col_authors = 'Author(s)'
        self.xls_col_repository = 'Repository'

        self.xls_col_duplicate_type = 'Duplicate Type'
        self.xls_val_by_doi = 'By DOI'
        self.xls_val_by_title = 'By Title'

        self.xls_columns = [self.xls_col_item,
                            self.xls_col_title,
                            self.xls_col_year,
                            self.xls_col_doi,
                            self.xls_col_document_type,
                            self.xls_col_languaje,
                            self.xls_col_cited_by,
                            self.xls_col_authors,
                            self.xls_col_repository]

        # Crossref API
        self.crossref_title = 'container-title'
        self.crossref_cited_by = 'is-referenced-by-count'
        self.crossref_created = 'created'
        self.crossref_created_date_parts = 'date-parts'
        self.crossref_type = 'type'
        self.crossref_language = 'language'

        # Status DOI
        self.status_inactive_doi = 'Inactive DOIs'

        # Fonts
        self.RED = '\033[31m'
        self.GREEN = '\033[32m'
        self.YELLOW = '\033[33m'
        self.BIRED = '\033[1;91m'
        self.BIGREEN = '\033[1;92m'
        self.END = '\033[0m'

    def show_print(self, message, logs = None, showdate = True, font = None, end = None):
        msg_print = message
        msg_write = message

        if font:
            msg_print = "%s%s%s" % (font, msg_print, self.END)

        if showdate is True:
            _time = time.strftime('%Y-%m-%d %H:%M:%S')
            msg_print = "%s %s" % (_time, msg_print)
            msg_write = "%s %s" % (_time, message)

        print(msg_print, end = end)
        if logs:
            for log in logs:
                if log:
                    with open(log, 'a', encoding = 'utf-8') as f:
                        f.write("%s\n" % msg_write)
                        f.close()

    def start_time(self):
        return time.time()

    def finish_time(self, start, message = None):
        finish = time.time()
        runtime = time.strftime("%H:%M:%S", time.gmtime(finish - start))
        if message is None:
            return runtime
        else:
            return "%s: %s" % (message, runtime)

    def create_directory(self, path):
        output = True
        try:
            if len(path) > 0 and not os.path.exists(path):
                os.makedirs(path)
        except Exception as e:
            output = False
        return output

    def check_path(self, path):
        _check = False
        if path:
            if len(path) > 0 and os.path.exists(path):
                _check = True
        return _check

    def remove_endpoint(self, text):
        _text = text.strip()

        while(_text[-1] == '.'):
            _text = _text[0:len(_text) - 1]
            _text = _text.strip()

        return _text

    def check_doi(self, doi):
        try:
            works = Works()
            response = works.doi(doi)

            is_valid = False
            status = None
            if response:
                status = response[self.crossref_title][0]
                if status != self.status_inactive_doi:
                    is_valid = True

            return is_valid
        except Exception as e:
            return False

    def get_language(self, code):
        hash_data = {
            'ab': 'Abkhazian',
            'aa': 'Afar',
            'af': 'Afrikaans',
            'ak': 'Akan',
            'sq': 'Albanian',
            'am': 'Amharic',
            'ar': 'Arabic',
            'an': 'Aragonese',
            'hy': 'Armenian',
            'as': 'Assamese',
            'av': 'Avaric',
            'ae': 'Avestan',
            'ay': 'Aymara',
            'az': 'Azerbaijani',
            'bm': 'Bambara',
            'ba': 'Bashkir',
            'eu': 'Basque',
            'be': 'Belarusian',
            'bn': 'Bengali',
            'bi': 'Bislama',
            'bs': 'Bosnian',
            'br': 'Breton',
            'bg': 'Bulgarian',
            'my': 'Burmese',
            'ca': 'Catalan',
            'km': 'Central Khmer',
            'ch': 'Chamorro',
            'ce': 'Chechen',
            'zh': 'Chinese',
            'cu': 'Church Slavic',
            'cv': 'Chuvash',
            'kw': 'Cornish',
            'co': 'Corsican',
            'cr': 'Cree',
            'hr': 'Croatian',
            'cs': 'Czech',
            'da': 'Danish',
            'dv': 'Divehi',
            'nl': 'Dutch',
            'dz': 'Dzongkha',
            'en': 'English',
            'eo': 'Esperanto',
            'et': 'Estonian',
            'ee': 'Ewe',
            'fo': 'Faroese',
            'fj': 'Fijian',
            'fi': 'Finnish',
            'fr': 'French',
            'ff': 'Fulah',
            'gd': 'Gaelic',
            'gl': 'Galician',
            'lg': 'Ganda',
            'ka': 'Georgian',
            'de': 'German',
            'el': 'Greek',
            'gn': 'Guarani',
            'gu': 'Gujarati',
            'ht': 'Haitian',
            'ha': 'Hausa',
            'he': 'Hebrew',
            'hz': 'Herero',
            'hi': 'Hindi',
            'ho': 'Hiri Motu',
            'hu': 'Hungarian',
            'is': 'Icelandic',
            'io': 'Ido',
            'ig': 'Igbo',
            'id': 'Indonesian',
            'ia': 'Interlingua',
            'ie': 'Interlingue',
            'iu': 'Inuktitut',
            'ik': 'Inupiaq',
            'ga': 'Irish',
            'it': 'Italian',
            'ja': 'Japanese',
            'jv': 'Javanese',
            'kl': 'Kalaallisut',
            'kn': 'Kannada',
            'kr': 'Kanuri',
            'ks': 'Kashmiri',
            'kk': 'Kazakh',
            'ki': 'Kikuyu',
            'rw': 'Kinyarwanda',
            'ky': 'Kirghiz',
            'kv': 'Komi',
            'kg': 'Kongo',
            'ko': 'Korean',
            'kj': 'Kuanyama',
            'ku': 'Kurdish',
            'lo': 'Lao',
            'la': 'Latin',
            'lv': 'Latvian',
            'li': 'Limburgan',
            'ln': 'Lingala',
            'lt': 'Lithuanian',
            'lu': 'Luba-Katanga',
            'lb': 'Luxembourgish',
            'mk': 'Macedonian',
            'mg': 'Malagasy',
            'ms': 'Malay',
            'ml': 'Malayalam',
            'mt': 'Maltese',
            'gv': 'Manx',
            'mi': 'Maori',
            'mr': 'Marathi',
            'mh': 'Marshallese',
            'mn': 'Mongolian',
            'na': 'Nauru',
            'nv': 'Navajo',
            'ng': 'Ndonga',
            'ne': 'Nepali',
            'nd': 'North Ndebele',
            'se': 'Northern Sami',
            'no': 'Norwegian',
            'nb': 'Norwegian Bokmal',
            'nn': 'Norwegian Nynorsk',
            'ny': 'Nyanja',
            'oc': 'Occitan',
            'oj': 'Ojibwa',
            'or': 'Oriya',
            'om': 'Oromo',
            'os': 'Ossetian',
            'pi': 'Pali',
            'ps': 'Pashto',
            'fa': 'Persian',
            'pl': 'Polish',
            'pt': 'Portuguese',
            'pa': 'Punjabi',
            'qu': 'Quechua',
            'ro': 'Romanian',
            'rm': 'Romansh',
            'rn': 'Rundi',
            'ru': 'Russian',
            'sm': 'Samoan',
            'sg': 'Sango',
            'sa': 'Sanskrit',
            'sc': 'Sardinian',
            'sr': 'Serbian',
            'sn': 'Shona',
            'ii': 'Sichuan Yi',
            'sd': 'Sindhi',
            'si': 'Sinhala',
            'sk': 'Slovak',
            'sl': 'Slovenian',
            'so': 'Somali',
            'nr': 'South Ndebele',
            'st': 'Southern Sotho',
            'es': 'Spanish',
            'su': 'Sundanese',
            'sw': 'Swahili',
            'ss': 'Swati',
            'sv': 'Swedish',
            'tl': 'Tagalog',
            'ty': 'Tahitian',
            'tg': 'Tajik',
            'ta': 'Tamil',
            'tt': 'Tatar',
            'te': 'Telugu',
            'th': 'Thai',
            'bo': 'Tibetan',
            'ti': 'Tigrinya',
            'to': 'Tonga',
            'ts': 'Tsonga',
            'tn': 'Tswana',
            'tr': 'Turkish',
            'tk': 'Turkmen',
            'tw': 'Twi',
            'ug': 'Uighur',
            'uk': 'Ukrainian',
            'ur': 'Urdu',
            'uz': 'Uzbek',
            've': 'Venda',
            'vi': 'Vietnamese',
            'vo': 'Volapük',
            'wa': 'Walloon',
            'cy': 'Welsh',
            'fy': 'Western Frisian',
            'wo': 'Wolof',
            'xh': 'Xhosa',
            'yi': 'Yiddish',
            'yo': 'Yoruba',
            'za': 'Zhuang',
            'zu': 'Zulu'
        }

        r = 'Unknown'
        if code in hash_data:
            r = hash_data[code]

        return r

    def get_document_type(self, code):
        # https://api.crossref.org/types
        hash_data = {
            'book': 'Book',
            'book-chapter': 'Book Chapter',
            'book-section': 'Book Section',
            'book-series': 'Book Series',
            'book-set': 'Book Set',
            'book-track': 'Book Track',
            'component': 'Component',
            'dataset': 'Dataset',
            'dissertation': 'Dissertation',
            'edited-book': 'Edited Book',
            'journal': 'Journal',
            'journal-article': 'Article',
            'journal-issue': 'Journal Issue',
            'journal-volume': 'Journal Volume',
            'monograph': 'Monograph',
            'other': 'Unknown Type',
            'book-part': 'Book Part',
            'peer-review': 'Review',
            'posted-content': 'Posted Content',
            'proceedings': 'Proceedings',
            'proceedings-article': 'Article; Proceedings Paper',
            'proceedings-series': 'Proceedings Series',
            'reference-book': 'Reference Book',
            'reference-entry': 'Reference Entry',
            'report': 'Report',
            'report-series': 'Report Series',
            'standard': 'Standard',
            'standard-series': 'Standard Series'
        }

        r = 'Unknown Type'
        if code in hash_data:
            r = hash_data[code]

        return r

    def get_complement(self, doi):
        try:
            works = Works()
            response = works.doi(doi)

            year = None
            cited_by = None
            language = None
            document_type = None
            if response:
                try:
                    year = response[self.crossref_created][self.crossref_created_date_parts][0][0]
                except Exception as e:
                    pass

                try:
                    cited_by = response[self.crossref_cited_by]
                except Exception as e:
                    pass

                try:
                    language = self.get_language(response[self.crossref_language])
                except Exception as e:
                    pass

                try:
                    document_type = self.get_document_type(response[self.crossref_type])
                except Exception as e:
                    pass

            return year, cited_by, language, document_type
        except Exception as e:
            return None, None, None, None

    def save_xls(self, dict_unique, dict_duplicates):

        def create_sheet(oworkbook, sheet_type, dictionary, styles_title, styles_rows):

            def add_row(pbar = None):
                icol = 0
                for irow, item in dictionary.items():
                    col_doi = item[self.xls_col_doi]
                    col_year = item[self.xls_col_year]
                    col_cited_by = item[self.xls_col_cited_by]
                    col_language = item[self.xls_col_languaje]
                    col_document_type = item[self.xls_col_document_type]

                    if pbar:
                        if col_year is None or col_cited_by is None or col_language is None or col_document_type is None:
                            _year, _cited_by, _language, _document_type = self.get_complement(col_doi)

                            col_year = _year if col_year is None else col_year
                            col_cited_by = _cited_by if col_cited_by is None else col_cited_by
                            col_language = _language if col_language is None else col_language
                            col_document_type = _document_type if col_document_type is None else col_document_type
                            pbar.update(1)

                    worksheet.write(irow, icol + 0, irow, styles_rows)
                    worksheet.write(irow, icol + 1, item[self.xls_col_title], styles_rows)
                    worksheet.write(irow, icol + 2, col_year, styles_rows)
                    worksheet.write(irow, icol + 3, col_doi, styles_rows)
                    worksheet.write(irow, icol + 4, col_document_type, styles_rows)
                    worksheet.write(irow, icol + 5, col_language, styles_rows)
                    worksheet.write(irow, icol + 6, col_cited_by, styles_rows)
                    worksheet.write(irow, icol + 7, item[self.xls_col_authors], styles_rows)
                    worksheet.write(irow, icol + 8, item[self.xls_col_repository], styles_rows)
                    if sheet_type == self.XLS_SHEET_DUPLICATES:
                        worksheet.write(irow, icol + 9, item[self.xls_col_duplicate_type], styles_rows)

            if sheet_type == self.XLS_SHEET_DUPLICATES:
                self.xls_columns.append(self.xls_col_duplicate_type)

            _last_col = len(self.xls_columns) - 1

            worksheet = oworkbook.add_worksheet(sheet_type)
            worksheet.freeze_panes(row = 1, col = 0) # Freeze the first row.
            worksheet.autofilter(first_row = 0, first_col = 0, last_row = 0, last_col = _last_col)
            worksheet.set_default_row(height = 14.5)

            # Add columns
            for icol, column in enumerate(self.xls_columns):
                worksheet.write(0, icol, column, styles_title)

            # Add rows
            worksheet.set_column(first_col = 0, last_col = 0, width = 7)  # Column A:A
            worksheet.set_column(first_col = 1, last_col = 1, width = 40) # Column B:B
            worksheet.set_column(first_col = 2, last_col = 2, width = 8)  # Column C:C
            worksheet.set_column(first_col = 3, last_col = 3, width = 33) # Column D:D
            worksheet.set_column(first_col = 4, last_col = 4, width = 18) # Column E:E
            worksheet.set_column(first_col = 5, last_col = 5, width = 12) # Column F:F
            worksheet.set_column(first_col = 6, last_col = 6, width = 11) # Column G:G
            worksheet.set_column(first_col = 7, last_col = 7, width = 36) # Column H:H
            worksheet.set_column(first_col = 8, last_col = 8, width = 13) # Column I:I
            if sheet_type == self.XLS_SHEET_DUPLICATES:
                worksheet.set_column(first_col = 9, last_col = 9, width = 19) # Column J:J

            total = 0
            for irow, item in dictionary.items():
                col_year = item[self.xls_col_year]
                col_cited_by = item[self.xls_col_cited_by]
                col_language = item[self.xls_col_languaje]
                col_document_type = item[self.xls_col_document_type]

                if col_year is None or col_cited_by is None or col_language is None or col_document_type is None:
                    total += 1

            if sheet_type == self.XLS_SHEET_DETAIL:
                with tqdm(total = total) as pbar:
                    add_row(pbar)
            elif sheet_type == self.XLS_SHEET_DUPLICATES:
                add_row()

        workbook = xlsxwriter.Workbook(self.XLS_FILE_OUTPUT)

        # Styles
        cell_format_title = workbook.add_format({'bold': True,
                                                 'font_color': 'white',
                                                 'bg_color': 'black',
                                                 'align': 'center',
                                                 'valign': 'vcenter'})
        cell_format_row = workbook.add_format({'text_wrap': True, 'valign': 'top'})

        self.show_print("Getting additional information from Crossref [Document Type, Language, Year, Cited by]", [self.LOG_FILE])
        create_sheet(workbook, self.XLS_SHEET_DETAIL, dict_unique, cell_format_title, cell_format_row)
        create_sheet(workbook, self.XLS_SHEET_DUPLICATES, dict_duplicates, cell_format_title, cell_format_row)
        self.show_print("", [self.LOG_FILE])

        workbook.close()

    def read_xls_summary(self, xlsfile):
        workbook = load_workbook(filename = xlsfile, data_only = True)
        sheet = workbook[self.XLS_SHEET_DETAIL]
        rows = sheet.rows

        file_collection = {}
        dois = []
        for index_i, row in enumerate(rows):
            doi = None
            collection = {}
            for index_j, cell in enumerate(row):
                if cell.value == self.xls_col_item:
                    break
                column_name = None
                if index_j == 0:
                    column_name = self.xls_col_item
                elif index_j == 1:
                    column_name = self.xls_col_title
                elif index_j == 2:
                    column_name = self.xls_col_year
                elif index_j == 3:
                    column_name = self.xls_col_doi
                    doi = cell.value
                    if doi:
                        dois.append(doi)
                elif index_j == 4:
                    column_name = self.xls_col_document_type
                elif index_j == 5:
                    column_name = self.xls_col_languaje
                elif index_j == 6:
                    column_name = self.xls_col_cited_by
                elif index_j == 7:
                    column_name = self.xls_col_authors

                _value = cell.value
                if column_name == self.xls_col_title:
                    if cell.value:
                        _value = self.remove_endpoint(cell.value)

                collection.update({column_name: _value})

            if len(collection) > 0:
                file_collection.update({index_i: collection})

        return file_collection, dois

    def get_list_files(self):
        if self.XLS_FILE_SCOPUS:
            self.DICT_XLS_FILES.update({self.REPOSITORY_SCOPUS: self.XLS_FILE_SCOPUS})
        if self.XLS_FILE_WOS:
            self.DICT_XLS_FILES.update({self.REPOSITORY_WOS: self.XLS_FILE_WOS})
        if self.XLS_FILE_PUBMED:
            self.DICT_XLS_FILES.update({self.REPOSITORY_PUBMED: self.XLS_FILE_PUBMED})
        if self.XLS_FILE_PUBMED_CENTRAL:
            self.DICT_XLS_FILES.update({self.REPOSITORY_PUBMED_CENTRAL: self.XLS_FILE_PUBMED_CENTRAL})
        if self.XLS_FILE_DIMENSIONS:
            self.DICT_XLS_FILES.update({self.REPOSITORY_DIMENSIONS: self.XLS_FILE_DIMENSIONS})

def main(args):
    try:
        start = orr.start_time()
        menu(args)

        orr.LOG_FILE = os.path.join(orr.OUTPUT_PATH, orr.LOG_NAME)
        orr.XLS_FILE_OUTPUT = os.path.join(orr.OUTPUT_PATH, orr.XLS_FILE_OUTPUT)
        orr.create_directory(orr.OUTPUT_PATH)
        orr.get_list_files()
        orr.show_print("#############################################################################", [orr.LOG_FILE], font = orr.BIGREEN)
        orr.show_print("############################# Remove Deplicates #############################", [orr.LOG_FILE], font = orr.BIGREEN)
        orr.show_print("#############################################################################", [orr.LOG_FILE], font = orr.BIGREEN)

        orr.show_print("Input files:", [orr.LOG_FILE], font = orr.GREEN)
        for index, (repository, file) in enumerate(orr.DICT_XLS_FILES.items()):
            if index == 0:
                base_repository = repository
                base_file = file
            orr.show_print("  %s" % file, [orr.LOG_FILE])
        orr.show_print("", [orr.LOG_FILE])

        del orr.DICT_XLS_FILES[base_repository]
        collection_base, dois_base = orr.read_xls_summary(base_file)
        for _, item in collection_base.items():
            item.update({orr.xls_col_repository: base_repository})

        collect_duplicate = {}
        collect_unique = {}
        for secondary_repository, secondary_file in orr.DICT_XLS_FILES.items():
            # Load information
            collection_secondary, dois_secondary = orr.read_xls_summary(secondary_file)

            # Get DOIs
            dois_duplicate = list(set(dois_base) & set(dois_secondary))
            dois_only_base = list(set(dois_base) - set(dois_secondary))
            dois_only_secondary = list(set(dois_secondary) - set(dois_base))

            # Get unique DOIs
            collect_unique_doi = {}
            index_u = 1
            for _, item in collection_base.items():
                doi = item[orr.xls_col_doi]
                if doi in dois_duplicate:
                    repository = '%s/%s' % (item[orr.xls_col_repository], secondary_repository)
                    item.update({orr.xls_col_repository: repository})

                collect_unique_doi.update({index_u: item})
                index_u += 1

            index_r = len(collect_duplicate) + 1
            index_u = len(collect_unique_doi) + 1
            for _, item in collection_secondary.items():
                doi = item[orr.xls_col_doi]
                item.update({orr.xls_col_repository: secondary_repository})
                if doi in dois_duplicate:
                    item[orr.xls_col_duplicate_type] = orr.xls_val_by_doi
                    collect_duplicate.update({index_r: item})
                    index_r += 1

                if doi in dois_only_secondary:
                    collect_unique_doi.update({index_u: item})
                    index_u += 1

            # Get duplicate titles
            nr_title = []
            re_title = []
            for _, row in collect_unique_doi.items():
                flag_unique = False

                title = row[orr.xls_col_title]
                if title:
                    title = title.strip().lower()
                    if title not in nr_title:
                        nr_title.append(title)
                        flag_unique = True

                    if not flag_unique:
                        re_title.append(title)

            # Get unique titles
            nr_title_ctrl = {item: {'n_check': 0, 'is_valid': False, 'repository': None} for item in re_title}
            index_u = 1
            index_r = len(collect_duplicate) + 1
            for _, row in collect_unique_doi.items():
                flag_unique = False

                doi = row[orr.xls_col_doi]
                title = row[orr.xls_col_title]
                if title:
                    title = title.strip().lower()

                if title in re_title:
                    _n_check = nr_title_ctrl[title]['n_check']
                    _is_valid = nr_title_ctrl[title]['is_valid']
                    _repository = nr_title_ctrl[title]['repository']

                    status = False
                    if not _is_valid:
                        status = orr.check_doi(doi)

                    if status:
                        flag_unique = True
                        nr_title_ctrl[title].update({'is_valid': True})
                    else:
                        if _n_check == 1 and _is_valid is False:
                            flag_unique = True # forced

                    if _repository is None:
                        _repository = row[orr.xls_col_repository]

                    if flag_unique:
                        row[orr.xls_col_repository] = '%s/%s' % (_repository, secondary_repository)
                    else:
                        row[orr.xls_col_repository] = row[orr.xls_col_repository].split('/')[-1] # secondary_repository

                    nr_title_ctrl[title].update({'n_check': _n_check + 1})
                    nr_title_ctrl[title].update({'repository': _repository})
                else:
                    flag_unique = True

                if flag_unique:
                    collect_unique.update({index_u: row})
                    index_u += 1
                else:
                    row[orr.xls_col_duplicate_type] = orr.xls_val_by_title
                    collect_duplicate.update({index_r: row})
                    index_r += 1

            # For other respositories
            collection_base = collect_unique.copy()
            dois_base = []
            for _, item in collection_base.items():
                dois_base.append(item[orr.xls_col_doi])

        # Create summary file
        orr.save_xls(collect_unique, collect_duplicate)
        orr.show_print("Output file: %s" % orr.XLS_FILE_OUTPUT, [orr.LOG_FILE], font = orr.GREEN)

        orr.show_print("", [orr.LOG_FILE])
        orr.show_print(orr.finish_time(start, "Elapsed time"), [orr.LOG_FILE])
        orr.show_print("Done!", [orr.LOG_FILE])
    except Exception as e:
        orr.show_print("\n%s" % traceback.format_exc(), [orr.LOG_FILE], font = orr.RED)
        orr.show_print(orr.finish_time(start, "Elapsed time"), [orr.LOG_FILE])
        orr.show_print("Done!", [orr.LOG_FILE])

if __name__ == '__main__':
    orr = RemoveDuplicate()
    main(sys.argv)
